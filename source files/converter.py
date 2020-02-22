from __future__ import annotations  # see https://www.python.org/dev/peps/pep-0563/
from typing import List, Set, Dict, Tuple, Optional, Union
from openpyxl import load_workbook
from collections import OrderedDict

import json
from os import path

HERE = path.dirname(path.abspath(__file__))
WORKBOOK = path.join(HERE, "OWC_Text.xlsx")


def main(workbook: str) -> None:
    """ Main app"""

    convert(workbook)


def convert(workbook: str) -> None:
    """ Convert workbook to JSON-LD in SKOS format """

    # load codes:
    codes = load_codes(workbook)

    # load template and transform into ordered dict:
    with open("template.json", 'r') as file:
        template = json.load(file)
    skos = OrderedDict()
    skos.update(template)

    concept_scheme = {"uri": "https://bartoc.org/owc/",
                      "type": "skos:ConceptScheme",
                      "label": "OWC Geographical Divisions"}

    graph = [concept_scheme]

    #
    wb = load_workbook(workbook)
    for ws in wb:
        for row in ws.iter_rows(min_row=3497, min_col=1, max_col=3, max_row=3508, values_only=True):  # max_row for dev

            entry = OrderedDict()

            # convert code to uri and add type (fixed):
            code = row[0]
            if code is None:
                continue
            code = code.replace(" ", "")
            uri = make_uri(code)
            entry.update({"uri": uri})
            entry.update({"type": "skos:Concept"})

            # add labels:
            labels = parse(row[1])
            entry.update(labels)

            # add hierarchy:
            entry.update(make_hierarchy(code, codes))

            graph.append(entry)

    skos.update({"graph": graph})

    # print for debug
    print(json.dumps(skos, indent=4, sort_keys=False))

    return skos


def load_codes(workbook) -> List[str]:
    """ Load all codes from workbook """

    codes = []
    wb = load_workbook(workbook)
    for ws in wb:
        for row in ws.iter_rows(min_row=7, min_col=1, max_col=1, values_only=True):
            code = row[0]
            if code is None:
                continue
            else:
                code = code.replace(" ", "")
                codes.append(code)
    return codes


def make_hierarchy(code: str, codes: List[str]) -> Dict:
    """ Add broader and narrower to concept """

    # top concept:
    if len(code) == 1:
        broader = None
        narrower = []
        for entry in codes:
            # exclude (other) top concepts (e.g., A):
            if len(entry) < 2:
                continue
            # first symbol of entry matches:
            if code[0] is entry[0]:
                # second symbol is number (e.g, A1):
                if entry[1] in str(set(range(0, 10))):
                    narrower.append({"uri": make_uri(entry)})
                # second symbol is not number and no other numbers (e.g., AA):
                elif len(entry) < 3:
                    narrower.append({"uri": make_uri(entry)})

    # middle concept:
    elif len(set(code).intersection(str(set(range(0, 10))))) == 0:
        print(f"middle concept with code {code}")
        broader = [{"uri": make_uri(code[0])}]
        narrower = []
        for entry in codes:
            if entry == code:
                continue
            elif code in entry and "." not in entry:
                narrower.append({"uri": make_uri(entry)})
            #elif entry.count(".") == 1 and code in ["OJ" + str(n) for n in range(6)]:
            #    narrower.append({"uri": make_uri(entry)})

        for thing in narrower:
            print(thing.values())


    # bottom concept:
    else:
        # OJ as special case:
        if "." in code:
            # bottom concept (e.g., OJ5.11.cAbau):
            if code.count(".") == 2:
                narrower = None
                broader = [{"uri": make_uri(code.split(".")[0] + "." + code.split(".")[1])}]
            # middle concept (e.g, OJ5.11):
            else:
                broader = [{"uri": make_uri(code.split(".")[0])}]
                narrower = []
                for entry in codes:
                    if (code in entry) and (code is not entry):
                        narrower.append({"uri": make_uri(entry)})
        else:
            narrower = None
            broader = [{"uri": make_uri(code[:2])}]

    hierarchy = {"broader": broader,
                 "narrower": narrower}

    return hierarchy


def make_uri(code: str) -> str:
    """ Convert OCW-code to URI """

    code = code.replace(" ", "")
    uri = "https://bartoc.org/ocw/" + code
    return uri


def parse(label: str) -> Dict:
    """ Parse OCW-label"""

    label_copy = label

    # prefLabel:
    label_split = label_copy.split(".", 1)
    value = label_split[0]
    pref_label = {"lang": "en",
                  "value": value}

    # definition:
    if len(label_split) > 1:
        value = label_split[1]
    else:
        value = None
    definition = {"lang": "en",
                  "value": value}  # text hinter dem punkt

    # altLabel:
    alt_label = None  # text in klammern oder k, evtl eher hiddenLabel

    # inScheme:
    in_scheme = None # wenn top label, dann name des schemas

    labels = {"prefLabel": pref_label,
              "altLabel": alt_label,
              "definition": definition,
              "inScheme": in_scheme}

    return labels

main(WORKBOOK)